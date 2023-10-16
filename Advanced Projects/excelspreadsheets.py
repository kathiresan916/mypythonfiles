"""Access Excel via Python"""

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the Excel workbook
wb = xl.load_workbook(
    r'C:\Users\HP\Desktop\My Knowledge Python\Advanced Projects\transactions.xlsx')

# Select the sheet named 'sheet1'
sheet = wb['Sheet1']

# Access the cell in the first row and first column
cell = sheet['A1']  # Using column letter 'A'
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value*0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                   min_col=4, max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save(r'C:\Users\HP\Desktop\My Knowledge Python\Advanced Projects\transactions1.xlsx')
